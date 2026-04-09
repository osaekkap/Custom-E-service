"""
Generate Swimlane Workflow Excel for Custom-E-service
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ──────────────────────────────────────────────────────────
C = {
    "title_bg":   "1F3864",  # dark navy
    "title_fg":   "FFFFFF",
    "lane_customer": "D6E4F0",   # light blue
    "lane_staff":    "D5F5E3",   # light green
    "lane_backend":  "FEF9E7",   # light yellow
    "lane_supabase": "F9EBEA",   # light pink
    "lane_nsw":      "F0ECF8",   # light purple
    "lane_header_customer": "2980B9",
    "lane_header_staff":    "27AE60",
    "lane_header_backend":  "F39C12",
    "lane_header_supabase": "E74C3C",
    "lane_header_nsw":      "8E44AD",
    "phase_bg":   "2C3E50",
    "phase_fg":   "FFFFFF",
    "step_arrow": "E8F6F3",
    "border_col": "BDC3C7",
    "white":      "FFFFFF",
    "check_green":"27AE60",
    "cross_red":  "E74C3C",
    "star_amber": "F39C12",
    "rbac_header":"1F3864",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin(color="BDC3C7"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium(color="7F8C8D"):
    s = Side(style="medium", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def apply_cell(ws, row, col, value="", bg=None, fg="000000",
               bold=False, h="center", v="center", wrap=True,
               size=10, italic=False, border=None):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size, italic=italic)
    c.alignment = align(h=h, v=v, wrap=wrap)
    if border:
        c.border = border
    return c

def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1,
                   end_row=r2, end_column=c2)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — SWIMLANE WORKFLOW
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Swimlane Workflow"
ws.sheet_view.showGridLines = False

# Column widths
col_widths = [18, 28, 28, 28, 22, 22]  # Phase | Customer | Staff | Backend | Supabase | NSW
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ── Title row ────────────────────────────────────────────────────────────
merge(ws, 1, 1, 2, 6)
apply_cell(ws, 1, 1,
           "Custom-E-service — System Swimlane Workflow",
           bg=C["title_bg"], fg=C["title_fg"],
           bold=True, size=16, border=border_medium())

# ── Lane headers ─────────────────────────────────────────────────────────
headers = [
    ("Phase", C["phase_bg"], C["phase_fg"]),
    ("Customer (Factory)\nCUSTOMER / CUST_ADMIN", C["lane_header_customer"], "FFFFFF"),
    ("NKTech Staff\nSTAFF / MANAGER / ADMIN", C["lane_header_staff"], "FFFFFF"),
    ("System (Backend API)\nNestJS", C["lane_header_backend"], "FFFFFF"),
    ("Supabase Auth\n+ Storage", C["lane_header_supabase"], "FFFFFF"),
    ("NSW / Customs\nกรมศุลกากร", C["lane_header_nsw"], "FFFFFF"),
]
for col, (label, bg, fg) in enumerate(headers, 1):
    apply_cell(ws, 3, col, label, bg=bg, fg=fg, bold=True, size=10,
               border=border_medium(C["border_col"]))
ws.row_dimensions[3].height = 40

# ── Helper: write a swimlane row ─────────────────────────────────────────
bg_map = [
    C["phase_bg"],
    C["lane_customer"],
    C["lane_staff"],
    C["lane_backend"],
    C["lane_supabase"],
    C["lane_nsw"],
]
fg_map = [
    C["phase_fg"], "000000", "000000", "000000", "000000", "000000"
]

def write_row(ws, row, cells, height=None, is_phase=False):
    """cells = list of 6 strings (one per lane)"""
    for col, val in enumerate(cells, 1):
        bg = C["phase_bg"] if (col == 1 and is_phase) else bg_map[col - 1]
        fg = C["phase_fg"] if (col == 1 and is_phase) else fg_map[col - 1]
        bold = is_phase and col == 1
        apply_cell(ws, row, col, val, bg=bg, fg=fg, bold=bold,
                   h="left" if col > 1 else "center",
                   border=border_thin(C["border_col"]))
    if height:
        ws.row_dimensions[row].height = height

# ── Phase rows ────────────────────────────────────────────────────────────
phases = [
    # (phase_label, [ [row cells…], … ], row_heights)
    (
        "① REGISTER\nB2B Onboarding",
        [
            ["", "กรอกฟอร์ม B2B\n• ชื่อบริษัท / เลขผู้เสียภาษี\n• ยอมรับ PDPA & T&C",
             "", "", "", ""],
            ["", "แนบเอกสาร\n• Company Certificate\n• PP20 (VAT)",
             "", "POST /auth/register/b2b\n↓ validate taxId unique\n↓ upload docs to Supabase Storage",
             "createUser()\nemail_confirm: true", ""],
            ["", "<- รับ companyCode\n<- Email confirmation",
             "<- Notification:\n'New B2B registration'\n(รอตรวจสอบ)",
             "↓ create Customer (TRIAL)\n+ Profile\n+ CustomerUser (ADMIN)", "", ""],
        ],
        [55, 55, 55],
    ),
    (
        "② LOGIN\nAuthentication",
        [
            ["", "กรอก email + password\n(หน้า Login)",
             "เหมือนกัน", "POST /auth/login", "signInWithPassword(email, pw)", ""],
            ["", "← access_token (JWT 15m)\n← refresh_token (7d)\n← user { role, customer }",
             "", "↓ check SUSPENDED?\n↓ sign JWT (role, customer_id)\n↓ create RefreshToken\n   (SHA-256 hash, 7d TTL)",
             "← OK / Error", ""],
        ],
        [45, 65],
    ),
    (
        "③ NEW SHIPMENT\nAI Path",
        [
            ["", "อัปโหลดไฟล์\n• Invoice (PDF/JPG)\n• Packing List\n• Booking\n• Privilege Docs",
             "", "POST /ai/extract-invoice\n↓ Claude Opus 4.6\n↓ extract fields", "", ""],
            ["", "← ผลลัพธ์ AI\n(confidence: high/med/low)\n↓ ตรวจสอบ / แก้ไข\n↓ ยืนยัน Submit",
             "", "↓ match HS codes\n↓ return extracted data", "", ""],
        ],
        [65, 65],
    ),
    (
        "③ NEW SHIPMENT\nManual Path",
        [
            ["", "กรอกฟอร์ม 5 sections:\n1. Document Control\n2. Exporter & Agent\n3. Invoice & Consignee\n4. Shipment Summary\n5. Line Items (HS Code)",
             "", "", "", ""],
            ["", "+ Privilege Flags\n(BOI / IEAT / FZ / 29BIS\n Re-Export / Re-Import)\n+ Upload Privilege Docs",
             "", "", "", ""],
            ["", "กด Submit →",
             "", "POST /jobs\n↓ create LogisticsJob (DRAFT)\n↓ POST /declarations\n   + DeclarationItems",
             "", ""],
            ["", "← jobNo: JOB-2026-XXXX",
             "← Notification:\n"New shipment created"", "", "", ""],
        ],
        [85, 55, 55, 45],
    ),
    (
        "④ APPROVAL\nWorkflow",
        [
            ["", "", "STAFF: ตรวจสอบเอกสาร\n↓ ขออนุมัติ",
             "PATCH /jobs/:id/request-approval\n↓ status: PENDING\n↓ create ApprovalLog", "", ""],
            ["", "", "MANAGER: รับ Notification\n↓ Approve / Reject",
             "PATCH /jobs/:id/approve\n↓ approvalStatus:\n   APPROVED / REJECTED", "", ""],
            ["", "← Notification:\n"Job approved / rejected"",
             "", "", "", ""],
        ],
        [55, 55, 35],
    ),
    (
        "⑤ NSW SUBMISSION\nebXML",
        [
            ["", "", "STAFF: กด Submit NSW",
             "POST /declarations/:id/submit-nsw\n↓ check approvalStatus == APPROVED\n↓ generate XML (XSD v4.00)",
             "", ""],
            ["", "", "",
             "↓ ebXML SOAP POST\n↓ retry: 1s → 2s → 4s (max 3)\n↓ status: SUBMITTING → SUBMITTED",
             "", "← NSW MSH endpoint\n← acknowledge\n← messageId"],
            ["", "", "",
             "↓ poll NSW status:\n   NSW_PROCESSING\n   → CUSTOMS_REVIEW\n   → CLEARED\n   → COMPLETED",
             "", "↓ PROCESSED\n↓ กรมศุลกากรตรวจ\n↓ CLEARED"],
            ["", "← Notification:\n"Cleared by customs"",
             "← Notification:\n"Job completed"", "", "", ""],
        ],
        [65, 65, 70, 40],
    ),
    (
        "⑥ BILLING\nInvoicing",
        [
            ["", "", "",
             "[auto] Job COMPLETED\n→ create BillingItem\n   (DECLARATION_FEE)", "", ""],
            ["", "", "NKTech สร้าง Invoice\nรวม unbilled items",
             "POST /billing/invoices\n↓ link BillingItems\n↓ status: SENT", "", ""],
            ["", "← Invoice PDF\n↓ ชำระเงิน",
             "อัปเดต status: PAID",
             "PATCH /billing/invoices/:id/status\n→ PAID", "", ""],
        ],
        [55, 55, 45],
    ),
    (
        "⑦ TOKEN REFRESH\nSilent Rotation",
        [
            ["", "[auto] access_token หมดอายุ\n→ HTTP 401\n→ axios interceptor trigger",
             "", "POST /auth/refresh\n↓ verify SHA-256 hash\n↓ check TTL & revoked",
             "", ""],
            ["", "← new access_token\n← new refresh_token\n↓ retry original request",
             "", "↓ revoke old token\n↓ issue new pair (rotate)", "", ""],
        ],
        [55, 45],
    ),
]

current_row = 4
for phase_label, rows, heights in phases:
    # Phase label cell — merged vertically
    total_rows = len(rows)
    merge(ws, current_row, 1, current_row + total_rows - 1, 1)
    apply_cell(ws, current_row, 1, phase_label,
               bg=C["phase_bg"], fg=C["phase_fg"],
               bold=True, size=10, h="center", v="center",
               border=border_medium())

    for i, (cells, h) in enumerate(zip(rows, heights)):
        r = current_row + i
        for col, val in enumerate(cells[1:], 2):  # skip col 1 (phase)
            bg = bg_map[col - 1]
            apply_cell(ws, r, col, val, bg=bg,
                       h="left", v="center",
                       border=border_thin(C["border_col"]))
        ws.row_dimensions[r].height = h

    current_row += total_rows

# ── Freeze panes ──────────────────────────────────────────────────────────
ws.freeze_panes = "B4"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RBAC PERMISSIONS
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("RBAC Permissions")
ws2.sheet_view.showGridLines = False

role_cols = ["SUPER_ADMIN", "TENANT_ADMIN", "MANAGER", "STAFF", "CUST_ADMIN", "CUSTOMER"]
role_colors = ["1F3864", "2980B9", "27AE60", "F39C12", "8E44AD", "E74C3C"]

permissions = [
    # (category, action, [values per role])
    # ✅ = full access, ❌ = no access, ✅* = scoped to own company
    ("Auth", "Login", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Auth", "Logout / Token Refresh", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Dashboard", "View Own Dashboard", ["✅", "✅", "✅", "✅", "✅*", "✅*"]),
    ("Dashboard", "View All Tenants (Super)", ["✅", "❌", "❌", "❌", "❌", "❌"]),
    ("Shipment", "Create New Shipment", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Shipment", "AI Extract Invoice", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Shipment", "View All Jobs (own tenant)", ["✅", "✅", "✅", "✅", "✅*", "✅*"]),
    ("Shipment", "Assign Job to Staff", ["✅", "✅", "✅", "✅", "❌", "❌"]),
    ("Approval", "Request Approval", ["✅", "✅", "✅", "✅", "❌", "❌"]),
    ("Approval", "Approve / Reject Job", ["✅", "✅", "✅", "❌", "❌", "❌"]),
    ("NSW", "Submit NSW Declaration", ["✅", "✅", "✅", "✅", "❌", "❌"]),
    ("NSW", "View Declaration Status", ["✅", "✅", "✅", "✅", "✅*", "✅*"]),
    ("Billing", "View Billing / Invoices", ["✅", "✅", "✅", "❌", "✅*", "✅*"]),
    ("Billing", "Create Invoice", ["✅", "✅", "✅", "❌", "❌", "❌"]),
    ("Billing", "Update Invoice Status", ["✅", "✅", "✅", "❌", "❌", "❌"]),
    ("Documents", "Upload Documents", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Documents", "View Documents", ["✅", "✅", "✅", "✅", "✅*", "✅*"]),
    ("Users", "Invite / Manage Users", ["✅", "✅", "❌", "❌", "✅*", "❌"]),
    ("Users", "Suspend User", ["✅", "✅", "❌", "❌", "❌", "❌"]),
    ("Master Data", "View HS Codes / Exporters", ["✅", "✅", "✅", "✅", "✅", "❌"]),
    ("Master Data", "Edit Master Data", ["✅", "✅", "✅", "✅", "❌", "❌"]),
    ("Notifications", "Receive Notifications", ["✅", "✅", "✅", "✅", "✅", "✅"]),
    ("Super Admin", "Super Admin Console", ["✅", "❌", "❌", "❌", "❌", "❌"]),
    ("Super Admin", "Manage All Tenants", ["✅", "❌", "❌", "❌", "❌", "❌"]),
    ("Super Admin", "System Settings", ["✅", "❌", "❌", "❌", "❌", "❌"]),
]

# Column widths
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 32
for i in range(len(role_cols)):
    ws2.column_dimensions[get_column_letter(i + 3)].width = 16

# Title
merge(ws2, 1, 1, 2, 8)
apply_cell(ws2, 1, 1,
           "Custom-E-service — RBAC Permission Matrix",
           bg=C["rbac_header"], fg="FFFFFF", bold=True, size=16,
           border=border_medium())

# Header row
apply_cell(ws2, 3, 1, "Category", bg=C["rbac_header"], fg="FFFFFF",
           bold=True, border=border_medium())
apply_cell(ws2, 3, 2, "Action / Permission", bg=C["rbac_header"], fg="FFFFFF",
           bold=True, border=border_medium())
ws2.row_dimensions[3].height = 35

for i, (role, color) in enumerate(zip(role_cols, role_colors)):
    apply_cell(ws2, 3, i + 3, role, bg=color, fg="FFFFFF",
               bold=True, size=10, border=border_medium())

# Data rows
cell_colors = {
    "✅": ("E8F8F5", "1E8449"),
    "✅*": ("EAF2FF", "1A5276"),
    "❌": ("FDEDEC", "922B21"),
}

prev_cat = None
row_num = 4
for cat, action, values in permissions:
    # Category alternating background
    cat_bg = "EBF5FB" if cat != prev_cat else "F4F6F7"
    prev_cat = cat

    apply_cell(ws2, row_num, 1, cat if cat != prev_cat else "",
               bg=cat_bg, bold=True, h="left",
               border=border_thin())
    # Always show category
    c = ws2.cell(row=row_num, column=1, value=cat)
    c.fill = fill(cat_bg)
    c.font = font(bold=True, size=9)
    c.alignment = align(h="left")
    c.border = border_thin()

    apply_cell(ws2, row_num, 2, action, bg="FDFEFE", h="left",
               border=border_thin())
    ws2.row_dimensions[row_num].height = 25

    for i, val in enumerate(values):
        bg, fg = cell_colors.get(val, ("FDFEFE", "000000"))
        apply_cell(ws2, row_num, i + 3, val, bg=bg, fg=fg,
                   bold=True, size=12, border=border_thin())

    row_num += 1

# Legend
row_num += 1
merge(ws2, row_num, 1, row_num, 2)
apply_cell(ws2, row_num, 1, "Legend:", bold=True, bg="F2F3F4",
           h="left", border=border_thin())
legends = [
    ("✅", "E8F8F5", "1E8449", "Full access"),
    ("✅*", "EAF2FF", "1A5276", "Scoped to own company only"),
    ("❌", "FDEDEC", "922B21", "No access"),
]
for i, (sym, bg, fg, label) in enumerate(legends):
    apply_cell(ws2, row_num, i + 3, f"{sym}  {label}", bg=bg, fg=fg,
               bold=False, h="left", border=border_thin())

ws2.freeze_panes = "C4"


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 — STATUS FLOW
# ════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Status Flow")
ws3.sheet_view.showGridLines = False

ws3.column_dimensions["A"].width = 20
ws3.column_dimensions["B"].width = 5
ws3.column_dimensions["C"].width = 20
ws3.column_dimensions["D"].width = 5
ws3.column_dimensions["E"].width = 20
ws3.column_dimensions["F"].width = 5
ws3.column_dimensions["G"].width = 20
ws3.column_dimensions["H"].width = 5
ws3.column_dimensions["I"].width = 20

merge(ws3, 1, 1, 2, 9)
apply_cell(ws3, 1, 1, "Custom-E-service — Status Flow Diagram",
           bg=C["title_bg"], fg="FFFFFF", bold=True, size=16,
           border=border_medium())

def status_box(ws, row, col, label, bg, fg="FFFFFF"):
    apply_cell(ws, row, col, label, bg=bg, fg=fg, bold=True,
               size=10, border=border_medium())
    ws.row_dimensions[row].height = 30

def arrow(ws, row, col, sym="→"):
    apply_cell(ws, row, col, sym, h="center", v="center", bold=True,
               size=14, fg="7F8C8D")

# ── Job Status Flow ──────────────────────────────────────────────────────
apply_cell(ws3, 3, 1, "Job Status Flow", bg="2C3E50", fg="FFFFFF",
           bold=True, border=border_medium())
ws3.row_dimensions[3].height = 28

job_statuses = [
    ("DRAFT", "7F8C8D"),
    ("PENDING", "F39C12"),
    ("IN_PROGRESS", "2980B9"),
    ("APPROVED", "27AE60"),
    ("SUBMITTING", "8E44AD"),
    ("SUBMITTED", "1ABC9C"),
    ("NSW_PROCESSING", "D35400"),
    ("CUSTOMS_REVIEW", "C0392B"),
    ("CLEARED", "196F3D"),
    ("COMPLETED", "1F3864"),
]
row = 4
for i, (status, color) in enumerate(job_statuses):
    col = (i * 2) + 1
    if col > 9:
        row += 2
        col = col - 9 if col <= 18 else col - 18
    status_box(ws3, row, col, status, color)
    if col < 9 and i < len(job_statuses) - 1 and (i + 1) % 5 != 0:
        arrow(ws3, row, col + 1)

# Add "REJECTED" branch
ws3.row_dimensions[6].height = 28
apply_cell(ws3, 6, 1, "⬇ REJECTED (from PENDING/IN_PROGRESS)",
           bg="E74C3C", fg="FFFFFF", bold=True, border=border_medium())

# ── Declaration Status Flow ──────────────────────────────────────────────
apply_cell(ws3, 8, 1, "Declaration Status Flow", bg="2C3E50", fg="FFFFFF",
           bold=True, border=border_medium())
ws3.row_dimensions[8].height = 28

decl_statuses = [
    ("DRAFT", "7F8C8D"),
    ("SUBMITTING", "8E44AD"),
    ("SUBMITTED", "1ABC9C"),
    ("NSW_PROCESSING", "D35400"),
    ("CUSTOMS_REVIEW", "C0392B"),
    ("CLEARED", "196F3D"),
    ("COMPLETED", "1F3864"),
    ("REJECTED", "E74C3C"),
    ("ERROR", "922B21"),
]
row = 9
for i, (status, color) in enumerate(decl_statuses[:7]):
    col = (i * 2) + 1
    status_box(ws3, row, col, status, color)
    if i < 6:
        arrow(ws3, row, col + 1)

row = 11
for i, (status, color) in enumerate(decl_statuses[7:]):
    col = (i * 2) + 1
    status_box(ws3, row, col, status, color)

# ── Invoice Status Flow ──────────────────────────────────────────────────
apply_cell(ws3, 13, 1, "Invoice Status Flow", bg="2C3E50", fg="FFFFFF",
           bold=True, border=border_medium())
ws3.row_dimensions[13].height = 28

invoice_statuses = [
    ("DRAFT", "7F8C8D"),
    ("SENT", "2980B9"),
    ("PAID", "27AE60"),
    ("OVERDUE", "E74C3C"),
    ("CANCELLED", "7F8C8D"),
]
row = 14
for i, (status, color) in enumerate(invoice_statuses):
    col = (i * 2) + 1
    status_box(ws3, row, col, status, color)
    if i < 3:
        arrow(ws3, row, col + 1)

ws3.freeze_panes = "A3"


# ── Save ────────────────────────────────────────────────────────────────────
output_path = r"C:\Users\Admin\Documents\NKTech\Custom-E-service\Custom-E-service_Swimlane_Workflow.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
