# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

def fill(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, color="000000", size=10): return Font(bold=bold, color=color, size=size, name="Arial")
def aln(h="center", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin(c="BDC3C7"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)
def med(c="7F8C8D"):
    s = Side(style="medium", color=c)
    return Border(left=s, right=s, top=s, bottom=s)

def wr(ws, row, col, val="", bg=None, fg="000000", bold=False,
       h="center", v="center", wrap=True, size=10, bdr=None):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = fill(bg)
    c.font = fnt(bold=bold, color=fg, size=size)
    c.alignment = aln(h=h, v=v, wrap=wrap)
    if bdr: c.border = bdr
    return c

def mg(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — SWIMLANE WORKFLOW
# ─────────────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Swimlane Workflow"
ws.sheet_view.showGridLines = False

COL_W = [22, 32, 32, 32, 26, 26]
LANE_BG = ["2C3E50", "D6E4F0", "D5F5E3", "FEF9E7", "F9EBEA", "F0ECF8"]
HDR_BG  = ["2C3E50", "2980B9", "27AE60", "E67E22", "E74C3C", "8E44AD"]

for i, w in enumerate(COL_W, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

mg(ws, 1, 1, 2, 6)
wr(ws, 1, 1, "Custom-E-service  \u2014  System Swimlane Workflow",
   bg="1F3864", fg="FFFFFF", bold=True, size=16, bdr=med())

HDR = [
    "Phase",
    "Customer (Factory)\nCUSTOMER / CUST_ADMIN",
    "NKTech Staff\nSTAFF / MANAGER / ADMIN",
    "System (Backend)\nNestJS API",
    "Supabase Auth\n+ Storage",
    "NSW / Customs\n(\u0e01\u0e23\u0e21\u0e28\u0e38\u0e25\u0e01\u0e32\u0e01\u0e23)",
]
for i, (h, bg) in enumerate(zip(HDR, HDR_BG), 1):
    wr(ws, 3, i, h, bg=bg, fg="FFFFFF", bold=True, size=10, bdr=med())
ws.row_dimensions[3].height = 44

PHASES = [
    (
        "\u2460 REGISTER\nB2B Onboarding",
        [
            ["", "\u0e01\u0e23\u0e2d\u0e01\u0e1f\u0e2d\u0e23\u0e4c\u0e21 B2B\n- \u0e0a\u0e37\u0e48\u0e2d\u0e1a\u0e23\u0e34\u0e29\u0e31\u0e17 / \u0e40\u0e25\u0e02\u0e1c\u0e39\u0e49\u0e40\u0e2a\u0e35\u0e22\u0e20\u0e32\u0e29\u0e35\n- \u0e22\u0e2d\u0e21\u0e23\u0e31\u0e1a PDPA & T&C", "", "", "", ""],
            ["", "\u0e41\u0e19\u0e1a\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23\n- Company Certificate (DBD)\n- PP20 (VAT Registration)", "", "POST /auth/register/b2b\n-> validate taxId unique\n-> upload docs to Storage", "createUser()\nemail_confirm: true", ""],
            ["", "<- \u0e23\u0e31\u0e1a companyCode\n<- Email confirmation", "<- Notification:\nNew B2B registration\n(\u0e23\u0e2d NKTech \u0e15\u0e23\u0e27\u0e08\u0e2a\u0e2d\u0e1a)", "-> create Customer (TRIAL)\n   + Profile\n   + CustomerUser (ADMIN)", "", ""],
        ],
        [55, 55, 55],
    ),
    (
        "\u2461 LOGIN\nAuthentication",
        [
            ["", "\u0e01\u0e23\u0e2d\u0e01 email + password", "\u0e40\u0e2b\u0e21\u0e37\u0e2d\u0e19\u0e01\u0e31\u0e19 (\u0e04\u0e19\u0e25\u0e30 role)", "POST /auth/login", "signInWithPassword(email, pw)", ""],
            ["", "<- access_token (JWT 15m)\n<- refresh_token (7d)\n<- user { role, customer_id }", "", "-> check SUSPENDED?\n-> sign JWT (role, customer_id)\n-> create RefreshToken\n   (SHA-256, 7d TTL)", "<- OK / Error", ""],
        ],
        [35, 65],
    ),
    (
        "\u2462 NEW SHIPMENT\n(AI Path)",
        [
            ["", "\u0e2d\u0e31\u0e1b\u0e42\u0e2b\u0e25\u0e14\u0e44\u0e1f\u0e25\u0e4c:\n- Invoice (PDF/JPG)\n- Packing List\n- Booking Confirmation\n- Privilege Documents", "", "POST /ai/extract-invoice\n-> Claude Opus 4.6\n-> extract all fields\n-> match HS codes", "", ""],
            ["", "<- \u0e1c\u0e25\u0e25\u0e31\u0e1e\u0e18\u0e4c AI\n   (confidence: high/med/low)\n-> \u0e15\u0e23\u0e27\u0e08\u0e2a\u0e2d\u0e1a / \u0e41\u0e01\u0e49\u0e44\u0e02\n-> \u0e22\u0e37\u0e19\u0e22\u0e31\u0e19 Submit", "", "<- return extracted data", "", ""],
        ],
        [75, 60],
    ),
    (
        "\u2462 NEW SHIPMENT\n(Manual Path)",
        [
            ["", "\u0e01\u0e23\u0e2d\u0e01\u0e1f\u0e2d\u0e23\u0e4c\u0e21 5 sections:\n1. Document Control\n2. Exporter & Agent\n3. Invoice & Consignee\n4. Shipment Summary\n5. Line Items (HS Code)", "", "", "", ""],
            ["", "+ Privilege Flags:\n   BOI / IEAT / FZ / 29BIS\n   Re-Export / Re-Import\n+ Upload Privilege Docs (PDF/JPG)", "", "", "", ""],
            ["", "\u0e01\u0e14 Submit ->", "", "POST /jobs\n-> create LogisticsJob (DRAFT)\nPOST /declarations\n-> create DeclarationItems", "", ""],
            ["", "<- jobNo: JOB-2026-XXXX", "<- Notification:\nNew shipment created", "", "", ""],
        ],
        [90, 60, 55, 40],
    ),
    (
        "\u2463 APPROVAL\nWorkflow",
        [
            ["", "", "STAFF: \u0e15\u0e23\u0e27\u0e08\u0e2a\u0e2d\u0e1a\u0e40\u0e2d\u0e01\u0e2a\u0e32\u0e23\n-> \u0e02\u0e2d\u0e2d\u0e19\u0e38\u0e21\u0e31\u0e15\u0e34", "PATCH /jobs/:id/request-approval\n-> status: PENDING\n-> create ApprovalLog", "", ""],
            ["", "", "MANAGER: \u0e23\u0e31\u0e1a Notification\n-> Approve / Reject", "PATCH /jobs/:id/approve\n-> approvalStatus:\n   APPROVED / REJECTED", "", ""],
            ["", "<- Notification:\nJob approved / rejected", "", "", "", ""],
        ],
        [55, 55, 35],
    ),
    (
        "\u2464 NSW SUBMISSION\nebXML v2.0",
        [
            ["", "", "STAFF: \u0e01\u0e14 Submit NSW", "POST /declarations/:id/submit-nsw\n-> check approvalStatus == APPROVED\n-> generate XML (XSD v4.00)", "", ""],
            ["", "", "", "-> ebXML SOAP POST\n-> retry: 1s -> 2s -> 4s (max 3)\n-> status: SUBMITTING -> SUBMITTED", "", "<- NSW MSH endpoint\n<- acknowledge + messageId"],
            ["", "", "", "-> poll NSW status:\n   NSW_PROCESSING\n   -> CUSTOMS_REVIEW\n   -> CLEARED\n   -> COMPLETED", "", "-> PROCESSED\n-> \u0e01\u0e23\u0e21\u0e28\u0e38\u0e25\u0e01\u0e32\u0e01\u0e23\u0e15\u0e23\u0e27\u0e08\u0e2a\u0e2d\u0e1a\n-> CLEARED"],
            ["", "<- Notification:\nCleared by customs", "<- Notification:\nJob completed", "", "", ""],
        ],
        [65, 65, 70, 40],
    ),
    (
        "\u2465 BILLING\nInvoicing",
        [
            ["", "", "", "[auto] Job COMPLETED\n-> create BillingItem\n   type: DECLARATION_FEE", "", ""],
            ["", "", "NKTech \u0e2a\u0e23\u0e49\u0e32\u0e07 Invoice\n(\u0e23\u0e27\u0e21 unbilled items)", "POST /billing/invoices\n-> link BillingItems\n-> status: SENT", "", ""],
            ["", "<- Invoice PDF\n-> \u0e0a\u0e33\u0e23\u0e30\u0e40\u0e07\u0e34\u0e19", "-> \u0e2d\u0e31\u0e1b\u0e40\u0e14\u0e15 status: PAID", "PATCH /billing/invoices/:id/status\n-> PAID", "", ""],
        ],
        [55, 55, 45],
    ),
    (
        "\u2466 TOKEN REFRESH\nSilent Rotation",
        [
            ["", "[auto] access_token \u0e2b\u0e21\u0e14\u0e2d\u0e32\u0e22\u0e38\n-> HTTP 401\n-> axios interceptor trigger", "", "POST /auth/refresh\n-> verify SHA-256 hash\n-> check TTL & not revoked", "", ""],
            ["", "<- new access_token\n<- new refresh_token\n-> retry original request", "", "-> revoke old token\n-> issue new token pair\n   (rotation)", "", ""],
        ],
        [60, 55],
    ),
]

cur = 4
for phase_lbl, rows, heights in PHASES:
    n = len(rows)
    mg(ws, cur, 1, cur + n - 1, 1)
    wr(ws, cur, 1, phase_lbl, bg="2C3E50", fg="FFFFFF", bold=True,
       h="center", v="center", bdr=med())
    for i, (vals, h) in enumerate(zip(rows, heights)):
        r = cur + i
        for col, v in enumerate(vals[1:], 2):
            wr(ws, r, col, v, bg=LANE_BG[col - 1], h="left", v="center", bdr=thin())
        ws.row_dimensions[r].height = h
    cur += n

ws.freeze_panes = "B4"

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — RBAC PERMISSIONS
# ─────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("RBAC Permissions")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 36
ROLES    = ["SUPER_ADMIN", "TENANT_ADMIN", "MANAGER", "STAFF", "CUST_ADMIN", "CUSTOMER"]
R_COLORS = ["1F3864", "2980B9", "27AE60", "E67E22", "8E44AD", "E74C3C"]
for i in range(6):
    ws2.column_dimensions[get_column_letter(i + 3)].width = 16

mg(ws2, 1, 1, 2, 8)
wr(ws2, 1, 1, "Custom-E-service  \u2014  RBAC Permission Matrix",
   bg="1F3864", fg="FFFFFF", bold=True, size=16, bdr=med())

wr(ws2, 3, 1, "Category",           bg="1F3864", fg="FFFFFF", bold=True, bdr=med())
wr(ws2, 3, 2, "Action / Permission", bg="1F3864", fg="FFFFFF", bold=True, bdr=med())
ws2.row_dimensions[3].height = 36
for i, (r, c) in enumerate(zip(ROLES, R_COLORS), 3):
    wr(ws2, 3, i, r, bg=c, fg="FFFFFF", bold=True, size=10, bdr=med())

PERMS = [
    ("Auth",        "Login",                     ["F","F","F","F","F","F"]),
    ("Auth",        "Logout / Token Refresh",     ["F","F","F","F","F","F"]),
    ("Dashboard",   "View Own Dashboard",         ["F","F","F","F","S","S"]),
    ("Dashboard",   "View All Tenants",           ["F","N","N","N","N","N"]),
    ("Shipment",    "Create New Shipment",        ["F","F","F","F","F","F"]),
    ("Shipment",    "AI Extract Invoice",         ["F","F","F","F","F","F"]),
    ("Shipment",    "View All Jobs",              ["F","F","F","F","S","S"]),
    ("Shipment",    "Assign Job to Staff",        ["F","F","F","F","N","N"]),
    ("Approval",    "Request Approval",           ["F","F","F","F","N","N"]),
    ("Approval",    "Approve / Reject Job",       ["F","F","F","N","N","N"]),
    ("NSW",         "Submit NSW Declaration",     ["F","F","F","F","N","N"]),
    ("NSW",         "View Declaration Status",    ["F","F","F","F","S","S"]),
    ("Billing",     "View Billing / Invoices",    ["F","F","F","N","S","S"]),
    ("Billing",     "Create Invoice",             ["F","F","F","N","N","N"]),
    ("Billing",     "Update Invoice Status",      ["F","F","F","N","N","N"]),
    ("Documents",   "Upload Documents",           ["F","F","F","F","F","F"]),
    ("Documents",   "View Documents",             ["F","F","F","F","S","S"]),
    ("Users",       "Invite / Manage Users",      ["F","F","N","N","S","N"]),
    ("Users",       "Suspend User",               ["F","F","N","N","N","N"]),
    ("Master Data", "View HS Codes / Exporters",  ["F","F","F","F","F","N"]),
    ("Master Data", "Edit Master Data",           ["F","F","F","F","N","N"]),
    ("Notif.",      "Receive Notifications",      ["F","F","F","F","F","F"]),
    ("Super Admin", "Super Admin Console",        ["F","N","N","N","N","N"]),
    ("Super Admin", "Manage All Tenants",         ["F","N","N","N","N","N"]),
    ("Super Admin", "System Settings",            ["F","N","N","N","N","N"]),
]
DISP  = {"F": "\u2713  Full", "S": "\u2713  Scoped", "N": "\u2717  None"}
STYLE = {"F": ("E8F8F5", "1E8449"), "S": ("EAF2FF", "1A5276"), "N": ("FDEDEC", "922B21")}
cat_palette = ["EBF5FB","F4F6F7","FEF9E7","F0ECF8","FDF2E9","EAFAF1","F9EBEA","F4ECF7","FDFEFE"]
CAT_BG = {}
for i, cat in enumerate(list(dict.fromkeys(p[0] for p in PERMS))):
    CAT_BG[cat] = cat_palette[i % len(cat_palette)]

rn = 4
for cat, action, vals in PERMS:
    wr(ws2, rn, 1, cat,    bg=CAT_BG[cat], bold=True, h="left", bdr=thin())
    wr(ws2, rn, 2, action, bg="FDFEFE",               h="left", bdr=thin())
    ws2.row_dimensions[rn].height = 24
    for i, v in enumerate(vals):
        bg, fg = STYLE[v]
        wr(ws2, rn, i + 3, DISP[v], bg=bg, fg=fg, bold=True, size=10, bdr=thin())
    rn += 1

rn += 1
mg(ws2, rn, 1, rn, 2)
wr(ws2, rn, 1, "Legend:", bold=True, bg="F2F3F4", h="left", bdr=thin())
for i, (v, label) in enumerate([
    ("F", "\u2713 Full \u2014 Full access"),
    ("S", "\u2713 Scoped \u2014 Own company only"),
    ("N", "\u2717 None \u2014 No access"),
]):
    bg, fg = STYLE[v]
    wr(ws2, rn, i + 3, label, bg=bg, fg=fg, h="left", bdr=thin())

ws2.freeze_panes = "C4"

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 3 — STATUS FLOW
# ─────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Status Flow")
ws3.sheet_view.showGridLines = False
for ci in range(1, 14):
    ws3.column_dimensions[get_column_letter(ci)].width = 19 if ci % 2 == 1 else 4

mg(ws3, 1, 1, 2, 13)
wr(ws3, 1, 1, "Custom-E-service  \u2014  Status Flow Overview",
   bg="1F3864", fg="FFFFFF", bold=True, size=16, bdr=med())

def sbox(ws, row, col, lbl, bg, fg="FFFFFF"):
    wr(ws, row, col, lbl, bg=bg, fg=fg, bold=True, size=10, bdr=med())
    ws.row_dimensions[row].height = 30

def arr(ws, row, col):
    wr(ws, row, col, "\u2192", h="center", bold=True, size=14, fg="95A5A6")

wr(ws3, 3, 1, "Job Status", bg="2C3E50", fg="FFFFFF", bold=True, bdr=med())
ws3.row_dimensions[3].height = 26
JOB1 = [("DRAFT","7F8C8D"),("PENDING","F39C12"),("IN_PROGRESS","2980B9"),("APPROVED","27AE60"),("SUBMITTING","8E44AD")]
JOB2 = [("SUBMITTED","1ABC9C"),("NSW_PROCESSING","D35400"),("CUSTOMS_REVIEW","C0392B"),("CLEARED","196F3D"),("COMPLETED","1F3864")]
for i, (lbl, col) in enumerate(JOB1):
    sbox(ws3, 4, i*2+1, lbl, col)
    if i < 4: arr(ws3, 4, i*2+2)
for i, (lbl, col) in enumerate(JOB2):
    sbox(ws3, 5, i*2+1, lbl, col)
    if i < 4: arr(ws3, 5, i*2+2)
mg(ws3, 6, 1, 6, 7)
wr(ws3, 6, 1, "Branch: REJECTED  (from PENDING or IN_PROGRESS -> back to DRAFT)",
   bg="E74C3C", fg="FFFFFF", bold=True, bdr=med())
ws3.row_dimensions[6].height = 26

wr(ws3, 8, 1, "Declaration Status", bg="2C3E50", fg="FFFFFF", bold=True, bdr=med())
ws3.row_dimensions[8].height = 26
DECL = [("DRAFT","7F8C8D"),("SUBMITTING","8E44AD"),("SUBMITTED","1ABC9C"),
        ("NSW_PROCESSING","D35400"),("CUSTOMS_REVIEW","C0392B"),("CLEARED","196F3D"),("COMPLETED","1F3864")]
for i, (lbl, col) in enumerate(DECL):
    sbox(ws3, 9, i*2+1, lbl, col)
    if i < 6: arr(ws3, 9, i*2+2)
mg(ws3, 10, 1, 10, 7)
wr(ws3, 10, 1, "Branches: REJECTED | ERROR  (from any step)",
   bg="E74C3C", fg="FFFFFF", bold=True, bdr=med())
ws3.row_dimensions[10].height = 26

wr(ws3, 12, 1, "Invoice Status", bg="2C3E50", fg="FFFFFF", bold=True, bdr=med())
ws3.row_dimensions[12].height = 26
INV = [("DRAFT","7F8C8D"),("SENT","2980B9"),("PAID","27AE60"),("OVERDUE","E74C3C"),("CANCELLED","95A5A6")]
for i, (lbl, col) in enumerate(INV):
    sbox(ws3, 13, i*2+1, lbl, col)
    if i < 3: arr(ws3, 13, i*2+2)
mg(ws3, 14, 1, 14, 9)
wr(ws3, 14, 1, "Note: OVERDUE auto-set when dueDate < today and status != PAID",
   bg="FEF9E7", fg="7F6000", h="left", bdr=thin())
ws3.row_dimensions[14].height = 24

ws3.freeze_panes = "A3"

# ─────────────────────────────────────────────────────────────────────────────
# Save
# ─────────────────────────────────────────────────────────────────────────────
out = r"C:\Users\Admin\Documents\NKTech\Custom-E-service\Custom-E-service_Swimlane_Workflow.xlsx"
wb.save(out)
print("Saved:", out)
